"""Query-state + confidence tests (Must-Have 5 states, confidence bonus) and
multi-turn filter refinement (Must-Have 7, Conversation 2)."""
from __future__ import annotations

import pytest
from fastapi.testclient import TestClient

from app.llm.provider import RuleBasedProvider
from app.main import app


@pytest.fixture()
def client():
    from app.api.routes import conversation_store

    conversation_store._conversations.clear()
    return TestClient(app)


CONTEXT_LISTING = {
    "last_intent": "transaction_list",
    "last_metric": "transaction_amount",
    "last_filters": {"transaction_type": "debit"},
    "last_date_range": {"type": "calendar_month", "start": "2026-08-01",
                        "end": "2026-08-31", "label": "Aug 2026"},
}


# ---------------------------------------------------------------------------
# Explicit query states (supported / empty_data / ambiguous / unsupported /
# invalid) — every chat response carries a machine-readable status.
# ---------------------------------------------------------------------------

def test_supported_status_and_high_confidence(client):
    body = client.post("/api/chat", json={
        "question": "How much did I spend last month?"}).json()
    assert body["status"] == "supported"
    assert body["confidence"] == "high"
    assert body["confidence_basis"]
    assert body["meta"]["grounded"] is True


def test_empty_data_status_on_zero_match(client):
    """A valid question with no matching data → empty_data + no_matches, a
    real zero — NOT a fabricated figure and NOT an error."""
    body = client.post("/api/chat", json={
        "question": "How much did I spend in January 2020?"}).json()
    assert body["refusal"] is None
    assert body["status"] == "empty_data"
    assert body["confidence"] == "no_matches"
    assert "no" in body["answer"].lower()


def test_ambiguous_status(client):
    body = client.post("/api/chat", json={
        "question": "How much moved last month?"}).json()
    assert body["status"] == "ambiguous"
    assert body["confidence"] == "none"
    assert body["evidence"] is None


def test_unsupported_status(client):
    body = client.post("/api/chat", json={
        "question": "How much did we spend on employee salaries?"}).json()
    assert body["status"] == "unsupported"
    assert body["confidence"] == "none"
    assert body["refusal"]["reason"] == "unsupported_metric"


def test_invalid_status_via_direct_query(client):
    body = client.post("/api/query", json={
        "intent": "transaction_summary", "metric": "transaction_amount",
        "aggregation": "sum", "date_range": {"type": "bogus_range"}}).json()
    assert body["status"] in ("invalid", "unsupported")
    assert body["refusal"] is not None


# ---------------------------------------------------------------------------
# Multi-turn filter refinement (spec Conversation 2)
# ---------------------------------------------------------------------------

def test_refinement_inherits_listing_filters():
    r = RuleBasedProvider().understand("Only those above ₹50,000.", CONTEXT_LISTING)
    assert r.query is not None
    assert r.query["intent"] == "transaction_list"
    assert r.query["filters"]["min_amount"] == 50000
    assert r.query["filters"]["transaction_type"] == "debit"  # inherited
    assert r.query["date_range"]["label"] == "Aug 2026"       # inherited


def test_refinement_type_switch():
    r = RuleBasedProvider().understand("Just the credits.", CONTEXT_LISTING)
    assert r.query["filters"]["transaction_type"] == "credit"
    assert r.query["date_range"]["label"] == "Aug 2026"


def test_refinement_requires_backreference():
    # a question with a NEW subject must not be treated as a refinement
    r = RuleBasedProvider().understand(
        "Which bank holds the most money?", CONTEXT_LISTING)
    assert r.query["intent"] == "bank_balance"


def test_refinement_without_context_is_normal_question():
    r = RuleBasedProvider().understand("Only those above ₹50,000.", {})
    # no prior listing → falls through to the standard amount-threshold path
    assert r.query is not None or r.refusal_reason is not None


def test_refinement_end_to_end_through_api(client):
    r1 = client.post("/api/chat", json={
        "question": "Show my largest debit transactions.",
        "conversation_id": "refine-1"}).json()
    assert r1["status"] == "supported"
    r2 = client.post("/api/chat", json={
        "question": "Only those above ₹50,000.",
        "conversation_id": "refine-1"}).json()
    assert r2["refusal"] is None, r2["answer"]
    assert r2["query"]["filters"].get("min_amount") == 50000
    assert r2["query"]["filters"].get("transaction_type") == "debit"
    assert r2["status"] == "supported"


# ---------------------------------------------------------------------------
# Evidence export (bonus): verbatim, masked rows
# ---------------------------------------------------------------------------

def test_export_csv_matches_displayed_records(client):
    chat = client.post("/api/chat", json={
        "question": "Show my largest transactions."}).json()
    records = chat["evidence"]["records"]
    resp = client.post("/api/export/evidence", json={
        "rows": records, "format": "csv"})
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith("text/csv")
    lines = resp.text.strip().splitlines()
    assert lines[0].startswith("transaction_id")
    # exactly the displayed rows (+ header)
    assert len(lines) == len(records) + 1
    # masked account numbers in the export too
    if "account_number" in lines[0]:
        col = lines[0].split(",").index("account_number")
        for line in lines[1:]:
            assert "XXXXX" in line.split(",")[col]


def test_export_rejects_empty(client):
    resp = client.post("/api/export/evidence", json={"rows": []})
    assert resp.status_code == 422


# ---------------------------------------------------------------------------
# Excel export (Phase 1): same rows, same masking, same columns as CSV/UI
# ---------------------------------------------------------------------------

def test_export_excel_parity_with_csv_and_ui(client):
    chat = client.post("/api/chat", json={
        "question": "Show my largest transactions."}).json()
    records = chat["evidence"]["records"]

    csv_resp = client.post("/api/export/evidence", json={
        "rows": records, "format": "csv"})
    xl_resp = client.post("/api/export/evidence", json={
        "rows": records, "format": "excel"})
    assert xl_resp.status_code == 200
    assert xl_resp.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    from openpyxl import load_workbook
    import io

    wb = load_workbook(io.BytesIO(xl_resp.content))
    ws = wb.active
    xl_rows = list(ws.iter_rows(values_only=True))
    assert len(xl_rows) == len(records) + 1  # header + rows

    csv_lines = csv_resp.text.strip().splitlines()
    assert len(csv_lines) == len(xl_rows)  # UI == CSV == Excel (row count)

    # same columns, same order
    csv_header = csv_lines[0].split(",")
    assert list(xl_rows[0]) == csv_header

    # same values cell-for-cell (CSV parses to the same strings)
    import csv as _csv
    csv_rows = list(_csv.reader(csv_lines[1:]))
    for xl_row, csv_row in zip(xl_rows[1:], csv_rows):
        for xl_cell, csv_cell in zip(xl_row, csv_row):
            assert str(xl_cell) == csv_cell or (xl_cell is None and csv_cell == "")

    # masking preserved in the Excel file
    if "account_number" in csv_header:
        col = csv_header.index("account_number")
        for xl_row in xl_rows[1:]:
            assert str(xl_row[col]).startswith("XXXXX")


def test_export_excel_requires_openpyxl_headers(client):
    chat = client.post("/api/chat", json={
        "question": "Show my largest transactions."}).json()
    resp = client.post("/api/export/evidence", json={
        "rows": chat["evidence"]["records"], "format": "excel"})
    assert resp.headers.get("content-disposition", "").endswith('filename="artha_evidence.xlsx"')


# ---------------------------------------------------------------------------
# Phase 2 — full confidence taxonomy (high / limited / no_matches / none)
# ---------------------------------------------------------------------------

def test_confidence_limited_on_thin_data(client):
    """A valid query matching <5 records → 'limited', not 'high'."""
    body = client.post("/api/chat", json={
        "question": "Show transactions above 5000000."}).json()
    matched = body["evidence"]["how_calculated"]["records_matched"]
    if 0 < matched < 5:
        assert body["confidence"] == "limited"
        assert "few records" in body["confidence_basis"] or \
            "indicative" in body["confidence_basis"]
    else:
        # dataset-dependent fallback: category still consistent with count
        assert body["confidence"] in ("high", "limited", "no_matches")


def test_confidence_high_requires_evidence(client):
    body = client.post("/api/chat", json={
        "question": "How much did I spend last month?"}).json()
    assert body["confidence"] == "high"
    assert body["evidence"] is not None
    assert body["evidence"]["grounded"] is True


def test_confidence_none_on_all_refusal_paths(client):
    for question in ("How much did we spend on salaries?",
                     "How much moved last month?"):
        body = client.post("/api/chat", json={"question": question}).json()
        assert body["confidence"] == "none"
        assert body["evidence"] is None


def test_confidence_never_probabilistic(client):
    """The confidence value must be one of the deterministic categories."""
    for question in ("What is my total available balance?",
                     "Show my largest transactions.",
                     "How much did I spend in January 2020?"):
        body = client.post("/api/chat", json={"question": question}).json()
        assert body["confidence"] in ("high", "limited", "no_matches", "none")
        assert isinstance(body["confidence_basis"], str)
